import os
import requests
import urllib.parse
import pythoncom
#from datetime import datetime, date, timedelta # anable if dont work
from flask import redirect, render_template, request, session
from functools import wraps
from openpyxl import load_workbook
import win32com.client

def apology(html, message, code=400):
    """Render message as an apology to user."""
    return render_template(html, code=code, message=message)


def login_required(f):
    """
    Decorate routes to require login.

    https://flask.palletsprojects.com/en/1.1.x/patterns/viewdecorators/
    """
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get("user_id") is None:
            return redirect("/login")
        return f(*args, **kwargs)
    return decorated_function


def render_xlsx(x, p):
    # insert data to xlsx form
    wb = load_workbook('static/ex_table.xlsx')
    sheets = wb.sheetnames
    sheet = wb[sheets[0]]
    sheet['G1'] = x["date"][:10]
    sheet['G3'] = x["install_date"]
    sheet['F1'] = "Offer №" + x["from_ord_numbs"] + " fr"
    sheet['C10'] = x['from_door_tipes']
    sheet['C11'] = x['from_st_size']
    sheet['C12'] = x['from_open_sides']
    sheet['C13'] = x['from_handle_colors']
    sheet['C14'] = x['from_paint_colors_frame_color'] + " + " + x['from_paint_colors_quoter_color']
    sheet['C16'] = x['from_covers_decor_color']
    sheet['C17'] = x['from_models_out_side_model'] + " / " + x['from_covers_out_side_cover']
    sheet['C18'] = x['from_models_in_side_model'] + " / " + x['from_covers_in_side_cover']
    sheet['C19'] = x['from_up_locks']
    sheet['C20'] = x['from_strips_up_lock_strip']
    sheet['C21'] = x['from_main_locks']
    sheet['C22'] = x['from_strips_main_lock_strip']
    sheet['C23'] = x['from_handles']
    sheet['C24'] = x['from_lock_cylinders']
    sheet['C25'] = x['from_peepholes']
    sheet['C26'] = x['from_latches']
    sheet['C38'] = "${:,.2f}".format(x['price'])
    sheet['G32'] = "${:,.2f}".format(x['deinstalling'])
    sheet['G34'] = "${:,.2f}".format(x['installing'])
    sheet['G36'] = "${:,.2f}".format(x['service_cost'])
    sheet['F37'] = "TOTAL: " + "${:,.2f}".format(x['total_price'])
    sheet['G10'] = "${:,.2f}".format(p[0])
    sheet['G26'] = "${:,.2f}".format(p[1])
    sheet['G17'] = "${:,.2f}".format(p[2] + p[3])
    sheet['G18'] = "${:,.2f}".format(p[4] + p[5])
    wb.save('static/ex_table.xlsx')


def render_pdf(new_pdf_name):
    # MAKE PDF FOR PRINT
    pythoncom.CoInitialize()
    o = win32com.client.Dispatch("Excel.Application")
    pythoncom.CoInitialize()
    # don't open window
    o.Visible = False
    # open work book
    wb_path = r'{}\static\ex_table.xlsx'.format(os.getcwd())
    wb = o.Workbooks.Open(wb_path)
    # open work Sheet1
    ws_index = 1
    ws = wb.Worksheets(ws_index)
    # add parameters for printing (not necessary)
    ws.PageSetup.Zoom = False
    ws.PageSetup.FitToPagesTall = 1
    ws.PageSetup.FitToPagesWide = 1
    ws.PageSetup.PrintArea = 'B1:P43'
    # create a new pdf in archive
    new_pdf_name = new_pdf_name + '.pdf'
    path_to_pdf = r'{0}\static\pdf_arch\{1}'.format(os.getcwd(), new_pdf_name)
    wb.WorkSheets(ws_index).Select()
    wb.ActiveSheet.ExportAsFixedFormat(0, path_to_pdf)
    # drop print area if used erlier
    ws.PageSetup.PrintArea = ""
    # close wb (important)
    wb.Close(True)
